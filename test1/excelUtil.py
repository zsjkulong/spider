import openpyxl
import datetime
import operator
from openpyxl.styles import colors
from openpyxl.styles import Font, Color, Fill,PatternFill
from openpyxl.cell import Cell
from test1.tushareUtil import AnalysisIndexData
import configparser
from test1.Item.indexItem import indexItem
class excelUtil:
    sheetName= ['每日指数','每日领涨概念板块','每日领涨行业板块','每日走势分析']
    path = ''
    numberOfGet = 5;
    yesterdayNumber = numberOfGet*1;
    beforeYesterdayNumber = numberOfGet*2;
    #exit = ['BK0815','BK0816']
    color = 'FFEEE8AA';

    #startRows = 0;

    def topN(self,item,startRow):#设置item的值到excel中
        #print(item);
        wb = openpyxl.load_workbook(self.path)
        if(operator.eq(item['isConcept'],'true')):
            sheet = wb.get_sheet_by_name(self.sheetName[1]);
        else:
            sheet = wb.get_sheet_by_name(self.sheetName[2]);
        sheet['B'+str(startRow)] = item['title']
        sheet['C' + str(startRow)] = item['code']
        self.setValAndFontColor(sheet['D' + str(startRow)],item['upRate'])
        self.setValAndFontColor(sheet['E' + str(startRow)], item['amount'])
        sheet['F' + str(startRow)] = item['ytitle']
        sheet['G' + str(startRow)] = item['ycode']
        self.setValAndFontColor(sheet['H' + str(startRow)],item['yupRate']);
        self.setValAndFontColor(sheet['I' + str(startRow)],item['yamount']);
        sheet['J' + str(startRow)] = item['bytitle']
        sheet['K' + str(startRow)] = item['bycode']
        self.setValAndFontColor(sheet['L' + str(startRow)],item['byupRate'])
        self.setValAndFontColor(sheet['M' + str(startRow)],item['byamount'])
        wb.save(self.path)
        #print("写入数据成功！")


    def makeSheetAndHeader(self):#写excel的header
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title =self.sheetName[0]
        sheet.merge_cells('B1:D1')
        sheet['B1'] = '上证指数';
        sheet['B2'] = '涨幅';
        sheet['C2'] = '成交量(亿)';
        sheet['D2'] = '成交量相比昨日';
        sheet.merge_cells('E1:G1')
        sheet['E1'] = '深证成指';
        sheet['E2'] = '涨幅';
        sheet['F2'] = '成交量(亿)';
        sheet['G2'] = '成交量相比昨日';

        sheet.merge_cells('H1:J1')
        sheet['H1'] = '创业板指';
        sheet['H2'] = '涨幅';
        sheet['I2'] = '成交量(亿)';
        sheet['J2'] = '成交量相比昨日';

        sheet.merge_cells('K1:M1')
        sheet['K1'] = '上证50';
        sheet['K2'] = '涨幅';
        sheet['L2'] = '成交量(亿)';
        sheet['M2'] = '成交量相比昨日';

        sheet.merge_cells('N1:P1')
        sheet['N1'] = '中小板指';
        sheet['N2'] = '涨幅';
        sheet['O2'] = '成交量(亿)';
        sheet['P2'] = '成交量相比昨日';
        sheet['A2'] = '日期'
        wb.save(self.path)

        #wb = openpyxl.Workbook()
        self.topNHeader(wb,1);
        self.topNHeader(wb,2)
        self.writeAnalysisHeader();



    def topNHeader(self,wb,i):
        sheet = wb.create_sheet(title=self.sheetName[i])
        # sheet.title = self.sheetName[1]
        sheet.merge_cells('B1:E1')
        sheet['B1'] = '当日领涨板块';
        sheet['B2'] = '板块名称';
        sheet['C2'] = '板块编码';
        sheet['D2'] = '涨幅';
        sheet['E2'] = '主力金额'
        sheet.merge_cells('F1:I1')
        sheet['F1'] = '昨日领涨板块';
        sheet['F2'] = '板块名称';
        sheet['G2'] = '板块编码';
        sheet['H2'] = '涨幅';
        sheet['I2'] = '主力金额'
        sheet.merge_cells('J1:M1')
        sheet['J1'] = '前日领涨板块';
        sheet['J2'] = '板块名称';
        sheet['K2'] = '板块编码';
        sheet['L2'] = '涨幅';
        sheet['M2'] = '主力金额'
        sheet['A2'] = '日期'
        wb.save(self.path)



    def readExcelRows(self,i):#获取excel的行数
        wb = openpyxl.load_workbook(self.path)
        sheet = wb.get_sheet_by_name(self.sheetName[i])
        return sheet.max_row+1;
        # sheet.get

    def readYesterday(self,startRowin,i):#读取昨天数据
        wb = openpyxl.load_workbook(self.path)
        sheet = wb.get_sheet_by_name(self.sheetName[i])
        if(startRowin and startRowin>0):
            startRow = startRowin;
        else:
            startRow = sheet.max_row+1
        list = [];
        if(startRow<=self.yesterdayNumber):
            return list;
        for cell in sheet['C'+str(startRow-self.yesterdayNumber):'C'+str(startRow-1)]:
            #print(cell[0].value);
            list.append(cell[0].value)
        print(list);
        return list;

    def readBeforeYesterday(self,startRowin,i):#读取前天数据
        wb = openpyxl.load_workbook(self.path)
        sheet = wb.get_sheet_by_name(self.sheetName[i])
        if (startRowin and startRowin > 0):
            startRow = startRowin;
        else:
            startRow = sheet.max_row + 1
        list = [];
        if (startRow <= self.beforeYesterdayNumber):
            return list;
        for cell in sheet['C' + str(startRow - self.beforeYesterdayNumber):'C' + str(startRow - 1 - self.yesterdayNumber)]:
            # print(cell[0].value);
            list.append(cell[0].value)
        #print(list);
        return list;

    def writerDate(self):#写日期
        self.wDate_two(self.sheetName[0])
        self.wDate_one(self.sheetName[1])
        self.wDate_one(self.sheetName[2])
        self.wDate_two(self.sheetName[3])



    def wDate_one(self,sheet):
        wb = openpyxl.load_workbook(self.path)
        sheet = wb.get_sheet_by_name(sheet)
        startRow = sheet.max_row + 1;
        mergeCell = 'A' + str(startRow - self.numberOfGet) + ':A' + str(startRow - 1);
        sheet.merge_cells(mergeCell);
        sheet['A' + str(startRow - self.numberOfGet)] = datetime.datetime.now().strftime('%Y-%m-%d')
        wb.save(self.path)

    def wDate_two(self,sheet):
        wb = openpyxl.load_workbook(self.path)
        sheet = wb.get_sheet_by_name(sheet)
        startRow = sheet.max_row;

        sheet['A' + str(startRow)] = datetime.datetime.now().strftime('%Y-%m-%d')
        wb.save(self.path)

    def setValAndFontColor(self,cell,val):#根据数字正负设置红绿
        if(operator.eq(val,'')):
            cell.value = val;
            return;

        if(float(val) > 0):
            ft = Font(color=colors.RED)
            cell.font = ft;
        elif(float(val) < 0):
            ft = Font(color=colors.GREEN)
            cell.font = ft;
        else:
            ft = Font(color=colors.BLACK)
            cell.font = ft;
        cell.value = val;


    def setCellColor(self):#设置隔一天数据的颜色

        self.setCellColorTool_one(self.sheetName[1],'A','M')

        self.setCellColorTool_two(self.sheetName[0], 'A', 'P')

        self.setCellColorTool_one(self.sheetName[2], 'A', 'M')

        self.setCellColorTool_two(self.sheetName[3], 'A', 'AE')

    def setCellColorTool_one(self,sheetName,start,end):
        wb = openpyxl.load_workbook(self.path)
        sheet = wb.get_sheet_by_name(sheetName)
        startRow = sheet.max_row + 1;
        # i = 0
        if (startRow % 2 == 0):
            for cells in sheet[(start + str(startRow - self.numberOfGet)):(end + str(startRow - 1))]:
                for cell in cells:
                    cell.fill = PatternFill(fill_type='solid', fgColor=self.color)
        wb.save(self.path);

    def setCellColorTool_two(self,sheetName,start,end):
        wb = openpyxl.load_workbook(self.path)
        sheet = wb.get_sheet_by_name(sheetName)
        startRow = sheet.max_row;
        # i = 0
        if (startRow % 2 == 0):
            for cells in sheet[start + str(startRow):end + str(startRow)]:
                for cell in cells:
                    cell.fill = PatternFill(fill_type='solid', fgColor=self.color)
                    # i += 1;

        wb.save(self.path);

    def makeItRed(self):#标记重复上榜的数据为红色或者紫色
        self.toRed(1);
        self.toRed(2);

    def toRed(self,i):
        wb = openpyxl.load_workbook(self.path)
        sheet = wb.get_sheet_by_name(self.sheetName[i])
        startRow = sheet.max_row + 1;
        yesterday = self.readYesterday(startRow - self.numberOfGet,i);
        beforeYesterday = self.readBeforeYesterday(startRow - self.numberOfGet,i);
        if (len(yesterday) == 0):
            return;
        if (len(beforeYesterday) == 0):
            return;

        i = self.numberOfGet;
        for cells in sheet['C' + str(startRow - self.numberOfGet):'C' + str(startRow - 1)]:
            if (cells[0].value in yesterday and cells[0].value in beforeYesterday):
                ft = Font(color=colors.RED)
                sheet['B' + str(startRow - i)].font = ft;
            elif (cells[0].value in yesterday or cells[0].value in beforeYesterday):
                ft = Font(color='FF9932CC')
                sheet['B' + str(startRow - i)].font = ft;
            else:
                i -= 1;
                continue;
            i -= 1;
        wb.save(self.path);

    def writeIndexData(self,indexItem):
        wb = openpyxl.load_workbook(self.path)
        sheet = wb.get_sheet_by_name(self.sheetName[0])

        startRow = sheet.max_row + 1;
        self.setRedOrGreen(sheet['B'+str(startRow)],indexItem.shrate);
        sheet['C' + str(startRow)].value = str(indexItem.shamount) ;
        if (startRow-1 <= 2):
            print()
        else:
            value = float(indexItem.shamount) - float(sheet['C' + str(startRow - 1)].value);
            self.setRedOrGreen(sheet['D' + str(startRow)], value);


        self.setRedOrGreen(sheet['E' + str(startRow)], indexItem.szrate);
        sheet['F' + str(startRow)].value = str(indexItem.szamount);
        if (startRow-1 <= 2):
            print()
        else:
            value = float(indexItem.szamount) - float(sheet['F' + str(startRow - 1)].value);
            self.setRedOrGreen(sheet['G' + str(startRow)], value);

        self.setRedOrGreen(sheet['H' + str(startRow)], indexItem.cyrate);
        sheet['I' + str(startRow)] = str(indexItem.cyamount);
        if (startRow-1 <= 2):
            print()
        else:
            value = float(indexItem.cyamount) - float(sheet['I' + str(startRow - 1)].value);
            self.setRedOrGreen(sheet['J' + str(startRow)], value);

        self.setRedOrGreen(sheet['K' + str(startRow)], indexItem.sh50rate);
        sheet['L' + str(startRow)] = str(indexItem.sh50amount);
        if (startRow-1 <= 2):
            print()
        else:
            value = float(indexItem.sh50amount) - float(sheet['L' + str(startRow - 1)].value);
            self.setRedOrGreen(sheet['M' + str(startRow)], value);

        self.setRedOrGreen(sheet['N' + str(startRow)], indexItem.zxrate);
        sheet['O' + str(startRow)] = str(indexItem.zxamount);
        if (startRow-1 <= 2):
            print()
        else:
            value = float(indexItem.zxamount) - float(sheet['O' + str(startRow - 1)].value);
            self.setRedOrGreen(sheet['P' + str(startRow)], value);


        wb.save(self.path);



    def setRedOrGreen(self,cell,value):
        ftRed = Font(color=colors.RED)
        ftGreen = Font(color=colors.GREEN)
        cell.value = value
        if(float(value) > 0):
            cell.font = ftRed;
        else:
            cell.font = ftGreen;

    def setCellRedOrGreen(self,cell,value):
        ftRed = Font(color=colors.RED)
        ftGreen = Font(color=colors.GREEN)
        ftBlue = Font(color=colors.BLUE)
        val = str(value).split('|');
        # cell.value = value
        if(len(val)>1):
            if(operator.eq(val[1],'red')):
                cell.font = ftRed;
            elif(operator.eq(val[1],'green')):
                cell.font = ftGreen;
            else:
                cell.font = ftBlue;
            cell.value = val[0];
        else:
            cell.value = value;

    def readConfig(self):
        cf = configparser.ConfigParser()
        cf.read('config.ini',encoding="utf-8-sig")
        self.numberOfGet = cf.getint('config', 'numberOfGet')
        self.path = cf.get('config', 'path')


    def writeIndexAnalysisData(self,AnalysisIndexData):
        wb = openpyxl.load_workbook(self.path)
        sheet = wb.get_sheet_by_name(self.sheetName[3])

        startRow = sheet.max_row + 1;
        self.setCellRedOrGreen(sheet['B' + str(startRow)], AnalysisIndexData.todayshClose)
        self.setCellRedOrGreen(sheet['C' + str(startRow)], AnalysisIndexData.shDire)
        self.setCellRedOrGreen(sheet['D' + str(startRow)], AnalysisIndexData.shValStatus)
        self.setCellRedOrGreen(sheet['E' + str(startRow)], AnalysisIndexData.shStatusHit)
        self.setCellRedOrGreen(sheet['F' + str(startRow)], AnalysisIndexData.shMACDHit)
        self.setCellRedOrGreen(sheet['G' + str(startRow)], AnalysisIndexData.shMACDDayHit)
        self.setCellRedOrGreen(sheet['H' + str(startRow)], AnalysisIndexData.todayszClose)
        self.setCellRedOrGreen(sheet['I' + str(startRow)], AnalysisIndexData.szDire)
        self.setCellRedOrGreen(sheet['J' + str(startRow)], AnalysisIndexData.szValStatus)
        self.setCellRedOrGreen(sheet['K' + str(startRow)], AnalysisIndexData.szStatusHit)
        self.setCellRedOrGreen(sheet['L' + str(startRow)], AnalysisIndexData.szMACDHit)
        self.setCellRedOrGreen(sheet['M' + str(startRow)], AnalysisIndexData.szMACDDayHit)
        self.setCellRedOrGreen(sheet['N' + str(startRow)], AnalysisIndexData.todaycyClose)
        self.setCellRedOrGreen(sheet['O' + str(startRow)], AnalysisIndexData.cyDire)
        self.setCellRedOrGreen(sheet['P' + str(startRow)], AnalysisIndexData.cyValStatus)
        self.setCellRedOrGreen(sheet['Q' + str(startRow)], AnalysisIndexData.cyStatusHit)
        self.setCellRedOrGreen(sheet['R' + str(startRow)], AnalysisIndexData.cyMACDHit)
        self.setCellRedOrGreen(sheet['S' + str(startRow)], AnalysisIndexData.cyMACDDayHit)

        self.setCellRedOrGreen(sheet['T' + str(startRow)], AnalysisIndexData.todaysh50Close)
        self.setCellRedOrGreen(sheet['U' + str(startRow)], AnalysisIndexData.sh50Dire)
        self.setCellRedOrGreen(sheet['V' + str(startRow)], AnalysisIndexData.sh50ValStatus)
        self.setCellRedOrGreen(sheet['W' + str(startRow)], AnalysisIndexData.sh50StatusHit)
        self.setCellRedOrGreen(sheet['X' + str(startRow)], AnalysisIndexData.sh50MACDHit)
        self.setCellRedOrGreen(sheet['Y' + str(startRow)], AnalysisIndexData.sh50MACDDayHit)

        self.setCellRedOrGreen(sheet['Z' + str(startRow)], AnalysisIndexData.todayzxClose)
        self.setCellRedOrGreen(sheet['AA' + str(startRow)], AnalysisIndexData.zxDire)
        self.setCellRedOrGreen(sheet['AB' + str(startRow)], AnalysisIndexData.zxValStatus)
        self.setCellRedOrGreen(sheet['AC' + str(startRow)], AnalysisIndexData.zxStatusHit)
        self.setCellRedOrGreen(sheet['AD' + str(startRow)], AnalysisIndexData.zxMACDHit)
        self.setCellRedOrGreen(sheet['AE' + str(startRow)], AnalysisIndexData.zxMACDDayHit)
        wb.save(self.path)


    def writeAnalysisHeader(self):
        wb = openpyxl.load_workbook(self.path)
        sheet = wb.create_sheet(title=self.sheetName[3])
        sheet.merge_cells('B1:G1')
        sheet['B1'] = '上证指数';
        sheet['B2'] = '指数';
        sheet['C2'] = '走势';
        sheet['D2'] = '成交量状态';
        sheet['E2'] = '走势成交量提示'
        sheet['F2'] = 'MACD提示'
        sheet['G2'] = '每日指数提示'
        sheet.merge_cells('H1:M1')
        sheet['H1'] = '深证成指';
        sheet['H2'] = '指数';
        sheet['I2'] = '走势';
        sheet['J2'] = '成交量状态';
        sheet['K2'] = '走势成交量提示'
        sheet['L2'] = 'MACD提示'
        sheet['M2'] = '每日指数提示'
        sheet.merge_cells('N1:S1')
        sheet['N1'] = '创业板指';
        sheet['N2'] = '指数';
        sheet['O2'] = '走势';
        sheet['P2'] = '成交量状态';
        sheet['Q2'] = '走势成交量提示'
        sheet['R2'] = 'MACD提示'
        sheet['S2'] = '每日指数提示'

        sheet.merge_cells('T1:Y1')
        sheet['T1'] = '上证50';
        sheet['T2'] = '指数';
        sheet['U2'] = '走势';
        sheet['V2'] = '成交量状态';
        sheet['W2'] = '走势成交量提示'
        sheet['X2'] = 'MACD提示'
        sheet['Y2'] = '每日指数提示'
        sheet.merge_cells('Z1:AE1')
        sheet['Z1'] = '中小板指';
        sheet['Z2'] = '指数';
        sheet['AA2'] = '走势';
        sheet['AB2'] = '成交量状态';
        sheet['AC2'] = '走势成交量提示'
        sheet['AD2'] = 'MACD提示'
        sheet['AE2'] = '每日指数提示'
        wb.save(self.path)
